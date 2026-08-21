# ====================
# Import external dependencies
# ====================
import csv, pytest, shutil
from openpyxl import load_workbook

# ====================
# Import internal dependencies
# ====================
from evaluator.commands.plot.plot import run_plot

pytestmark = pytest.mark.skipif(
    shutil.which('Rscript') is None,
    reason='Rscript not on PATH',
)

# ====================
# Define constants
# ====================
ANALYSE_COLUMNS = [
    'tomogram', 'label', 'equiv_diameter_nm', 'major_axis_diameter', 'minor_axis_diameter',
    'aspect_ratio', 'eccentricity', 'membrane_volume', 'lumen_volume', 'surface_area',
    'is_enclosed', 'closure_fill_ratio',
]

# ====================
# Define helper functions
# ====================
def _write_analyse_csv(path, n=12):
    with path.open('w', newline='') as fh:
        writer = csv.writer(fh)
        writer.writerow(ANALYSE_COLUMNS)
        for i in range(n):
            writer.writerow([
                'tomo1.mrc', i, 60.0 + i, 70.0 + i, 50.0 + i, 1.2, 0.4,
                40000 + i * 100, 25000 + i * 50, 18000 + i * 80,
                bool(i % 3), 0.02 + 0.01 * i,
            ])

# ====================
# Define module fixtures
# ====================
@pytest.fixture(scope='module')
def analyse_csv(tmp_path_factory):
    path = tmp_path_factory.mktemp('plot_data') / 'analyse.csv'
    _write_analyse_csv(path)
    return path

@pytest.fixture(scope='module')
def plot_output(analyse_csv, tmp_path_factory):
    out = tmp_path_factory.mktemp('plot_output')
    run_plot(
        analyse_csv, None, out,
        sections=['distributions', 'qc'], all_sections=False,
        overwrite=False, rscript=None,
    )
    return out / 'evaluator' / 'plot'

# ====================
# Define tests
# ====================
class TestPlotDistributionsQC:
    def test_distributions_summary_stats_written(self, plot_output):
        path = plot_output / 'distributions' / 'summary_stats.xlsx'
        assert path.exists()
        wb = load_workbook(path)
        assert wb.active.max_row > 1  # header + at least one feature row

    def test_qc_summary_written(self, plot_output):
        path = plot_output / 'qc' / 'qc_summary.xlsx'
        assert path.exists()
        wb = load_workbook(path)
        row = list(wb.active.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        assert row[0] == 'tomo1.mrc'  # tomogram column

    def test_density_svg_written_per_feature(self, plot_output):
        assert (plot_output / 'distributions' / 'equiv_diameter_nm_density.svg').exists()

    def test_index_md_lists_both_sections(self, plot_output):
        index = (plot_output / 'index.md').read_text()
        assert 'distributions/' in index and 'qc/' in index

class TestPlotConcordance:
    def test_concordance_requires_matching_rows(self, tmp_path, analyse_csv):
        model_json = tmp_path / 'model.json'
        model_json.write_text(
            "{'results': [{'label_id': 999, 'source_file': 'no_such_tomo.mrc', "
            "'chosen_model': 'sphere', 'radius': 40.0, 'rmse_nm': 1.0, "
            "'reliability': {'is_reliable': true}}]}"
        )
        out = tmp_path / 'out'
        run_plot(
            analyse_csv, model_json, out,
            sections=['concordance'], all_sections=False,
            overwrite=False, rscript=None,
        )
        section_dir = out / 'evaluator' / 'plot' / 'concordance'
        assert not (section_dir / 'concordance_stats.xlsx').exists()
